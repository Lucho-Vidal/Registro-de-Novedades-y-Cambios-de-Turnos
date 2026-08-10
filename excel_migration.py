"""Importación inicial desde las cuatro hojas operativas del Excel."""

from datetime import datetime
import re

import openpyxl


SHEETS = {
    "base": "BASE",
    "tipos": "TipoNovedad",
    "novedades": "NOVEDADES",
    "cambios": "Cambio de Turnos",
}


def _normalizar(valor):
    texto = str(valor or "").strip().upper()
    return re.sub(r"[^A-Z0-9]", "", texto)


def _texto(valor):
    return "" if valor is None else str(valor).strip()


def _entero(valor):
    try:
        return int(str(valor).strip())
    except (TypeError, ValueError):
        return None


def _headers(ws):
    return {_normalizar(cell.value): index for index, cell in enumerate(ws[1])}


def _valor(row, headers, *nombres):
    for nombre in nombres:
        index = headers.get(_normalizar(nombre))
        if index is not None and index < len(row):
            return row[index]
    return None


def _fecha(valor):
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.isoformat(timespec="seconds")
    return _texto(valor)


def migrate_workbook(workbook_path, store):
    """Importa datos sin tocar el XLSX. Devuelve cantidades por tabla."""
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    store.initialize()
    counts = {key: 0 for key in SHEETS}
    now = store.now()

    with store.write_transaction() as connection:
        ws = workbook[SHEETS["base"]]
        headers = _headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            legajo = _entero(_valor(row, headers, "LEGAJO SAP", "LEGAJO", "LEGAJOS SAP"))
            nombre = _texto(_valor(row, headers, "APELLIDOS Y NOMBRES", "APELLIDOS Y NOMBRES"))
            if legajo is None or not nombre:
                continue
            connection.execute(
                """INSERT INTO empleados
                   (legajo, apellidos_nombres, especialidad, dotacion, turnos, franco, actualizado_en)
                   VALUES (?, ?, ?, ?, ?, ?, ?)
                   ON CONFLICT(legajo) DO UPDATE SET
                     apellidos_nombres=excluded.apellidos_nombres,
                     especialidad=excluded.especialidad,
                     dotacion=excluded.dotacion,
                     turnos=excluded.turnos,
                     franco=excluded.franco,
                     actualizado_en=excluded.actualizado_en""",
                (
                    legajo,
                    nombre,
                    _texto(_valor(row, headers, "ESPECIALIDAD")),
                    _texto(_valor(row, headers, "DOTACION", "DOT")),
                    _texto(_valor(row, headers, "TURNOS", "TURNO")),
                    _texto(_valor(row, headers, "FRANCO", "DSO")),
                    now,
                ),
            )
            counts["base"] += 1

        ws = workbook[SHEETS["tipos"]]
        for row in ws.iter_rows(values_only=True):
            nombre = _texto(row[0] if row else None)
            if not nombre or _normalizar(nombre) in {"TIPONOVEDAD", "NOVEDAD"}:
                continue
            connection.execute("INSERT OR IGNORE INTO tipos_novedad(nombre) VALUES (?)", (nombre,))
            counts["tipos"] += 1

        ws = workbook[SHEETS["novedades"]]
        headers = _headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            record_id = _entero(_valor(row, headers, "ID"))
            legajo = _entero(_valor(row, headers, "LEGAJO SAP", "LEGAJO"))
            novedad = _texto(_valor(row, headers, "NOVEDAD"))
            if record_id is None or legajo is None or not novedad:
                continue
            connection.execute(
                """INSERT OR IGNORE INTO novedades
                   (id, registrado_en, legajo, apellidos_nombres, especialidad, dotacion,
                    turnos, franco, novedad, fecha_inicio, fecha_fin, referencia_estacion,
                    supervisor, observaciones, usuario_windows)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    record_id,
                    _texto(_valor(row, headers, "FECHA", "FECHA Y HORA")),
                    legajo,
                    _texto(_valor(row, headers, "APELLIDOS Y NOMBRES")),
                    _texto(_valor(row, headers, "ESPECIALIDAD")),
                    _texto(_valor(row, headers, "DOTACION", "DOT")),
                    _texto(_valor(row, headers, "TURNOS", "TURNO")),
                    _texto(_valor(row, headers, "FRANCO", "DSO")),
                    novedad,
                    _fecha(_valor(row, headers, "FECHA DE INICIO NOVEDAD")),
                    _fecha(_valor(row, headers, "FECHA DE FIN NOVEDAD")),
                    _texto(_valor(row, headers, "REFERENCIA ESTACIÓN", "REFERENCIA ESTACION")),
                    _texto(_valor(row, headers, "SUPERVISOR")),
                    _texto(_valor(row, headers, "OBSERVACIONES")),
                    _texto(_valor(row, headers, "USUARIO WINDOWS")),
                ),
            )
            counts["novedades"] += 1

        ws = workbook[SHEETS["cambios"]]
        headers = _headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            record_id = _entero(_valor(row, headers, "ID"))
            legajo_1 = _entero(_valor(row, headers, "LEGAJO"))
            legajo_2 = _entero(_valor(row, headers, "LEGAJO2"))
            if record_id is None or legajo_1 is None or legajo_2 is None:
                continue
            connection.execute(
                """INSERT OR IGNORE INTO cambios_turno
                   (id, registrado_en, legajo_1, apellidos_nombres_1, especialidad_1,
                    dotacion_1, turnos_1, franco_1, legajo_2, apellidos_nombres_2,
                    especialidad_2, dotacion_2, turnos_2, franco_2, fecha_cambio,
                    referencia_estacion, supervisor, observaciones, usuario_windows)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    record_id,
                    _texto(_valor(row, headers, "FECHA", "FECHA Y HORA")),
                    legajo_1,
                    _texto(_valor(row, headers, "APELLIDOS Y NOMBRES")),
                    _texto(_valor(row, headers, "ESPECIALIDAD")),
                    _texto(_valor(row, headers, "DOTACION")),
                    _texto(_valor(row, headers, "TURNOS")),
                    _texto(_valor(row, headers, "FRANCO")),
                    legajo_2,
                    _texto(_valor(row, headers, "APELLIDOS Y NOMBRES2")),
                    _texto(_valor(row, headers, "ESPECIALIDAD2")),
                    _texto(_valor(row, headers, "DOTACION2")),
                    _texto(_valor(row, headers, "TURNOS2")),
                    _texto(_valor(row, headers, "FRANCO2")),
                    _fecha(_valor(row, headers, "FECHA DE CAMBIO DE TURNO")),
                    _texto(_valor(row, headers, "REFERENCIA ESTACIÓN", "REFERENCIA ESTACION")),
                    _texto(_valor(row, headers, "SUPERVISOR")),
                    _texto(_valor(row, headers, "OBSERVACIONES")),
                    _texto(_valor(row, headers, "USUARIO WINDOWS")),
                ),
            )
            counts["cambios"] += 1
    workbook.close()
    return counts


def migrate_operational_sheet(workbook_path, store, sheet_name, clear_existing=False):
    """Importa solo NOVEDADES o Cambio de Turnos y conserva los datos existentes."""
    if sheet_name not in {SHEETS["novedades"], SHEETS["cambios"]}:
        raise ValueError("Solo se pueden importar NOVEDADES o Cambio de Turnos.")
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    store.initialize()
    count = 0
    with store.write_transaction() as connection:
        if clear_existing:
            table = "novedades" if sheet_name == SHEETS["novedades"] else "cambios_turno"
            connection.execute(f"DELETE FROM {table}")
        ws = workbook[sheet_name]
        headers = _headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            record_id = _entero(_valor(row, headers, "ID"))
            if sheet_name == SHEETS["novedades"]:
                legajo = _entero(_valor(row, headers, "LEGAJO SAP", "LEGAJO"))
                novedad = _texto(_valor(row, headers, "NOVEDAD"))
                if record_id is None or legajo is None or not novedad:
                    continue
                connection.execute(
                    """INSERT OR IGNORE INTO novedades
                       (id, registrado_en, legajo, apellidos_nombres, especialidad, dotacion,
                        turnos, franco, novedad, fecha_inicio, fecha_fin, referencia_estacion,
                        supervisor, observaciones, usuario_windows)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (record_id, _texto(_valor(row, headers, "FECHA", "FECHA Y HORA")), legajo,
                     _texto(_valor(row, headers, "APELLIDOS Y NOMBRES")), _texto(_valor(row, headers, "ESPECIALIDAD")),
                     _texto(_valor(row, headers, "DOTACION", "DOT")), _texto(_valor(row, headers, "TURNOS", "TURNO")),
                     _texto(_valor(row, headers, "FRANCO", "DSO")), novedad,
                     _fecha(_valor(row, headers, "FECHA DE INICIO NOVEDAD")), _fecha(_valor(row, headers, "FECHA DE FIN NOVEDAD")),
                     _texto(_valor(row, headers, "REFERENCIA ESTACIÓN", "REFERENCIA ESTACION")),
                     _texto(_valor(row, headers, "SUPERVISOR")), _texto(_valor(row, headers, "OBSERVACIONES")),
                     _texto(_valor(row, headers, "USUARIO WINDOWS"))),
                )
            else:
                legajo_1 = _entero(_valor(row, headers, "LEGAJO"))
                legajo_2 = _entero(_valor(row, headers, "LEGAJO2"))
                if record_id is None or legajo_1 is None or legajo_2 is None:
                    continue
                connection.execute(
                    """INSERT OR IGNORE INTO cambios_turno
                       (id, registrado_en, legajo_1, apellidos_nombres_1, especialidad_1,
                        dotacion_1, turnos_1, franco_1, legajo_2, apellidos_nombres_2,
                        especialidad_2, dotacion_2, turnos_2, franco_2, fecha_cambio,
                        referencia_estacion, supervisor, observaciones, usuario_windows)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (record_id, _texto(_valor(row, headers, "FECHA", "FECHA Y HORA")), legajo_1,
                     _texto(_valor(row, headers, "APELLIDOS Y NOMBRES")), _texto(_valor(row, headers, "ESPECIALIDAD")),
                     _texto(_valor(row, headers, "DOTACION")), _texto(_valor(row, headers, "TURNOS")), _texto(_valor(row, headers, "FRANCO")),
                     legajo_2, _texto(_valor(row, headers, "APELLIDOS Y NOMBRES2")), _texto(_valor(row, headers, "ESPECIALIDAD2")),
                     _texto(_valor(row, headers, "DOTACION2")), _texto(_valor(row, headers, "TURNOS2")), _texto(_valor(row, headers, "FRANCO2")),
                     _fecha(_valor(row, headers, "FECHA DE CAMBIO DE TURNO")),
                     _texto(_valor(row, headers, "REFERENCIA ESTACIÓN", "REFERENCIA ESTACION")),
                     _texto(_valor(row, headers, "SUPERVISOR")), _texto(_valor(row, headers, "OBSERVACIONES")),
                     _texto(_valor(row, headers, "USUARIO WINDOWS"))),
                )
            count += 1
    workbook.close()
    return count


def migrate_empleados_sheet(workbook_path, store, clear_existing=False):
    """Importa la hoja BASE usando el legajo como clave de actualización."""
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    store.initialize()
    count = 0
    with store.write_transaction() as connection:
        if clear_existing:
            connection.execute("DELETE FROM empleados")
        ws = workbook[SHEETS["base"]]
        headers = _headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            legajo = _entero(_valor(row, headers, "LEGAJO SAP", "LEGAJO", "LEGAJOS SAP"))
            nombre = _texto(_valor(row, headers, "APELLIDOS Y NOMBRES"))
            if legajo is None or not nombre:
                continue
            connection.execute(
                """INSERT INTO empleados
                   (legajo, apellidos_nombres, especialidad, dotacion, turnos, franco, actualizado_en)
                   VALUES (?, ?, ?, ?, ?, ?, ?)
                   ON CONFLICT(legajo) DO UPDATE SET
                     apellidos_nombres=excluded.apellidos_nombres,
                     especialidad=excluded.especialidad,
                     dotacion=excluded.dotacion,
                     turnos=excluded.turnos,
                     franco=excluded.franco,
                     activo=1,
                     actualizado_en=excluded.actualizado_en""",
                (legajo, nombre, _texto(_valor(row, headers, "ESPECIALIDAD")),
                 _texto(_valor(row, headers, "DOTACION", "DOT")),
                 _texto(_valor(row, headers, "TURNOS", "TURNO")),
                 _texto(_valor(row, headers, "FRANCO", "DSO")), store.now()),
            )
            count += 1
    workbook.close()
    return count
