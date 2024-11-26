import tkinter as tk
from tkinter import ttk
from tkinter import messagebox,scrolledtext 
from tkinter import filedialog 
from datetime import datetime
from ttkbootstrap import DateEntry
from ttkbootstrap import Style
from ttkbootstrap.widgets import Entry
from datetime import datetime, timedelta
from tkinter.scrolledtext import ScrolledText
import openpyxl
import os
import time
import threading
import ctypes
import math

class FormularioExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.geometry('1110x650')  # Ajuste inicial de la ventana
        root.state('zoomed')
        self.root.title("Registro de Novedades y Cambios de turnos TK")
        user32 = ctypes.windll.user32
        user32.SetProcessDPIAware()
        self.WIDTH, self.HEIGHT  = user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)
        self.WIDTH = math.floor(self.WIDTH * 0.99)
        self.HEIGHT = math.floor(self.HEIGHT * 0.78)

        self.labelCarga = tk.Label(self.root, text="Cargando Excel...")
        self.labelCarga.grid(row=1, column=0, padx=10, pady=0, sticky="w")

        self.leer_archivo_base()
        self.theme_file = 'theme'
        self.theme = self.cargar_tema()

        # Verificar si el archivo existe; si no, crearlo
        if not os.path.exists(self.excel_file):
            self.crear_archivo_excel()
        
        self.current_view = 'table'
        self.cargar_excel()
        
        
        # Cargar opciones de tipo de novedad
        self.tipo_novedades = []
        self.cargarTipoNovedades()
        
        # Aplicar el estilo ttkbootstrap
        self.style = Style()
        self.style.theme_use(self.theme)  # Tema inicial
        # Obtener temas disponibles (puedes definir manualmente los temas si es necesario)
        if hasattr(self.style, "theme_names"):
            self.temas = self.style.theme_names()
        else:
            # Lista predeterminada de temas en ttkbootstrap
            self.temas = [
                "cosmo", "litera", "minty", "pulse", "quartz",
                "flatly", "journal", "solar", "cerculean",
                "darkly", "sandstone", "superhero", "morph"
            ]
        
        # Configurar el menú principal
        self.menu_bar = tk.Menu(self.root)
        self.root.config(menu=self.menu_bar)

        # Menú Archivo
        self.archivo_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Archivo", menu=self.archivo_menu)

        # Opción para seleccionar archivo
        self.archivo_menu.add_command(label="Seleccionar archivo", command=self.seleccionar_archivo)
        # self.archivo_menu.add_command(label="Maximizar", command=self.cambiar_tamanio_ventana)
        
        # Menú "Opciones"
        self.opciones_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="Opciones", menu=self.opciones_menu)
        
        # Submenú "Seleccionar Tema"
        self.temas_menu = tk.Menu(self.opciones_menu, tearoff=0)
        self.opciones_menu.add_cascade(label="Seleccionar Tema", menu=self.temas_menu)
        
        # Agregar temas al submenú
        for tema in self.temas:
            self.temas_menu.add_command(label=tema, command=lambda t=tema: self.cambiar_tema(t))
        
        # Etiqueta de ejemplo para visualizar el tema
        self.label = tk.Label(self.root, text=f"Tema actual: {self.theme}", font=("Arial", 8))
        self.label.grid(row=1, column=0, padx=10, pady=0, sticky="e")
        

        # Variables del formulario
        self.legajo_var = tk.StringVar()
        self.apellidos_nombres_var = tk.StringVar()
        self.especialidad_var = tk.StringVar()
        self.dotacion_var = tk.StringVar()
        self.turnos_var = tk.StringVar()
        self.franco_var = tk.StringVar()
        self.novedad_var = tk.StringVar()
        self.fecha_inicio_novedad_var = tk.StringVar()
        self.fecha_fin_novedad_var = tk.StringVar()
        self.referencia_estacion_var = tk.StringVar()
        self.supervisor_var = tk.StringVar()
        self.observaciones_var = tk.StringVar()
        
        self.legajo_2_var = tk.StringVar()
        self.apellidos_nombres_2_var = tk.StringVar()
        self.especialidad_2_var = tk.StringVar()
        self.dotacion_2_var = tk.StringVar()
        self.turnos_2_var = tk.StringVar()
        self.franco_2_var = tk.StringVar()
        self.fecha_cambio_turno_var = tk.StringVar()

        # Marco principal
        main_frame = tk.Frame(self.root)
        main_frame.grid(row=0, column=0, sticky="nsew")

        # Crear marcos de formulario y tabla
        self.form_frame = tk.Frame(main_frame)
        self.form_cambios_frame = tk.Frame(main_frame)
        self.table_cambios_frame = tk.Frame(main_frame)
        self.table_frame = tk.Frame(main_frame)
        self.table_frame.grid(row=0, column=0, padx=10, pady=10)  # Muestra la tabla desde el inicio

        # Configurar el grid de la ventana principal para que el marco se expanda 
        self.root.grid_rowconfigure(0, weight=1) 
        self.root.grid_columnconfigure(0, weight=1)
        
        self.crear_tabla_novedades()

    def cargar_excel(self):
        try:
            print(f"Abriendo archivo Excel: {self.excel_file}")
            self.labelCarga.config(text="Actualizando....")
            self.wb = openpyxl.load_workbook(self.excel_file)
            self.sheet_base = self.wb["BASE"]
            self.sheet_novedades = self.wb["NOVEDADES"]
            self.sheet_tipo_novedad = self.wb["TipoNovedad"]
            self.sheet_cambio_turnos = self.wb["Cambio de Turnos"]

            self.actualizar_tabla()
            if(self.current_view == 'table'):
                print("Novedades actualizadas correctamente.")
                self.labelCarga.config(text="Novedades actualizadas correctamente.")
            elif(self.current_view == 'table_cambios'):
                print("Cambios de turnos actualizados correctamente.")
                self.labelCarga.config(text="Cambios de turnos actualizados correctamente.")
        except FileNotFoundError:
            print("El archivo Excel no se encontró.")
            self.labelCarga.config(text="Archivo Excel no encontrado.")
        except Exception as e:
            print(f"Error abriendo o cargando el archivo Excel: {e}")
            self.labelCarga.config(text="Error al cargar el archivo Excel.")
    
    # Función para actualizar la tabla con datos (ejemplo usando Listbox)
    def actualizar_tabla(self):
        if not (self.apellido_filter_var.get() == "Buscar por nombre") or not (self.dotacion_filter_var.get() == "Todas"):
            return
        if self.sheet_novedades  and self.current_view == 'table':
            # Limpiar la tabla antes de actualizar
            self.tabla_novedades.delete(*self.tabla_novedades.get_children())
            # for item in self.tabla_novedades.get_children():  # Iterar sobre los ítems existentes
            #     self.tabla_novedades.delete(item)  # Eliminar cada ítem
            # Leer las primeras filas de la hoja de "BASE" y mostrarlas en la tabla
            for row in self.sheet_novedades.iter_rows(min_row=2, values_only=True):  # min_row=2 para omitir encabezados
                # Reemplazar valores None por "-"
                row_data = ["-" if celda is None else celda for celda in row]
                self.tabla_novedades.insert("", "end", values=row_data)  # Insertar los datos en la tabla
        
        if self.sheet_cambio_turnos and self.current_view == "table_cambios":
            for item in self.table_cambios.get_children():  # Iterar sobre los ítems existentes
                self.table_cambios.delete(item)  # Eliminar cada ítem
            for row in self.sheet_cambio_turnos.iter_rows(min_row=2, values_only=True):  # min_row=2 para omitir encabezados
                # Reemplazar valores None por "-"
                row_data = ["-" if celda is None else celda for celda in row]
                self.table_cambios.insert("", "end", values=row_data) 

    # Función que ejecuta la tarea periódica
    def ejecutar_periodicamente(self):
        while True:
            self.cargar_excel()  # Ejecutar la lectura del archivo
            time.sleep(60)

    def leer_archivo_base(self):
        try:
            with open("path_base", "r", encoding="utf-8") as file:
                self.excel_file = file.read().strip()
                self.excel_file = self.excel_file.replace("\\", "\\\\")  # Reemplazar '\' por '\\'
        except Exception as e:
            print(f"Error leyendo el archivo: {e}")
            self.excel_file = r'PLANILLA NOVEDADES PERSONAL ABORDO.xlsx'

    def seleccionar_archivo(self):
        """Abre el explorador de archivos y guarda la ruta seleccionada en un archivo."""
        # Abrir el cuadro de diálogo para seleccionar archivo
        archivo_seleccionado = filedialog.askopenfilename(title="Seleccionar archivo", filetypes=[("Excel", "*.xlsx*")])

        if archivo_seleccionado:  # Si se seleccionó un archivo
            try:
                # Guardar la ruta en el archivo 'path_base'
                with open("path_base", "w", encoding="utf-8") as f:
                    f.write(archivo_seleccionado)
                print(f"Ruta guardada: {archivo_seleccionado}")
                self.excel_file = archivo_seleccionado
                self.cargar_excel()
                self.actualizar_tabla()
            except Exception as e:
                print(f"Error al guardar la ruta: {e}")

    def cambiar_tema(self, nuevo_tema):
        """Cambiar el tema y guardarlo en un archivo"""
        try:
            self.style.theme_use(nuevo_tema)
            self.theme = nuevo_tema
            with open(self.theme_file, "w", encoding="utf-8") as file:
                file.write(nuevo_tema)
            self.label.config(text=f"Tema actual: {self.theme}")
            print(f"Tema cambiado a: {nuevo_tema}")
        except Exception as e:
            print(f"Error al cambiar el tema: {e}")

    def cargar_tema(self):
        """Cargar el tema almacenado en el archivo"""
        try:
            with open(self.theme_file, "r", encoding="utf-8") as file:
                return file.read().strip()  # Leer y limpiar espacios
        except FileNotFoundError:
            # Si no existe el archivo, devolver un tema por defecto
            return "flatly"
        except Exception as e:
            print(f"Error al cargar el tema: {e}")
            return "flatly"  # Tema por defecto en caso de error
    def toggle_view(self, target_view=None):
        """
        Cambia a la vista especificada. Si no se proporciona 'target_view',
        alterna entre las vistas según el estado actual.
        """

        # Ocultar todas las vistas
        self.form_frame.grid_forget()
        self.table_frame.grid_forget()
        self.form_cambios_frame.grid_forget()
        self.table_cambios_frame.grid_forget()

        # Determinar la vista objetivo
        self.current_view = target_view  # Si se especifica un destino, cambiar a esa vista

        # Mostrar la vista correspondiente
        if self.current_view == "form":
            self.form_frame.grid(row=0, column=0, padx=10, pady=10)
            self.mostrar_formulario_novedades()
        elif self.current_view == "form_cambios":
            self.form_cambios_frame.grid(row=0, column=0, padx=10, pady=10)
            self.mostrar_formulario_cambios()
        elif self.current_view == "table_cambios":
            self.table_cambios_frame.grid(row=0, column=0, padx=10, pady=10)
            self.crear_tabla_cambios()
        else:
            self.table_frame.grid(row=0, column=0, padx=10, pady=10)
            self.crear_tabla_novedades()
    
    def cargarTipoNovedades(self):
        for row in self.sheet_tipo_novedad.iter_rows(min_row=2, values_only=True):
            tipo_novedad = row[0]  # Suponemos que los tipos de novedad están en la primera columna
            self.tipo_novedades.append(tipo_novedad)

    def crear_archivo_excel(self):
        """Crea un archivo Excel con encabezados si no existe"""
        wb = openpyxl.Workbook()
        
        # Crear la hoja "BASE" y agregar encabezados
        sheet_base = wb.create_sheet(title="BASE")  # Crea la hoja "BASE"
        encabezados_base = [
            "LEGAJO","APELLIDOS Y NOMBRES", "ESPECIALIDAD",
            "DOTACION", "TURNOS","FRANCO"
        ]
        sheet_base.append(encabezados_base)  # Agregar los encabezados a la hoja "BASE"
        # Crear la hoja "NOVEDADES" y agregar encabezados
        sheet_novedades = wb.create_sheet(title="NOVEDADES")  # Crea la hoja "NOVEDADES"

        encabezados_novedades = [
            "ID","Fecha y hora","LEGAJO ", "APELLIDOS Y NOMBRES", "ESPECIALIDAD",
            "DOTACION", "TURNOS","FRANCO", 
            "NOVEDAD", "Fecha de Inicio Novedad", "Fecha de Fin Novedad",
            "REFERENCIA ESTACIÓN", "SUPERVISOR", "Observaciones"
        ]
        sheet_novedades.append(encabezados_novedades)  # Agregar los encabezados a la hoja "NOVEDADES"
        
        # Crear la hoja "TipoNovedad" y agregar encabezados
        sheet_tiponovedad = wb.create_sheet(title="TipoNovedad")  # Crea la hoja "TipoNovedad"
        encabezados_tiponovedad = ["Enfermo"]
        sheet_tiponovedad.append(encabezados_tiponovedad)  # Agregar los encabezados a la hoja "TipoNovedad"
        
        # Crear la hoja "Cambio de Turnos" y agregar encabezados
        sheet_cambio_turnos = wb.create_sheet(title="Cambio de Turnos")  # Crea la hoja "Cambio de Turnos"

        encabezados_cambio_turnos = [
            "ID","Fecha y hora", "LEGAJO", "APELLIDOS Y NOMBRES", 
            "ESPECIALIDAD", "DOTACION", "TURNOS", "FRANCO", 
            "LEGAJO2", "APELLIDOS Y NOMBRES2", "ESPECIALIDAD2",
            "DOTACION2", "TURNOS2","FRANCO2", 
            "Fecha de Cambio de Turno","REFERENCIA ESTACIÓN", "SUPERVISOR", "Observaciones"
        ]
        sheet_cambio_turnos.append(encabezados_cambio_turnos)  # Agregar los encabezados a la hoja "Cambio de Turnos"
        
        # Eliminar la hoja predeterminada (por defecto, openpyxl crea una hoja vacía llamada "Sheet")
        del wb["Sheet"]
        
        # Guardar el archivo
        wb.save(self.excel_file)
        print(f"Archivo creado:{self.excel_file}")
        
    def cargar_datos_completos_novedades(self):
        try:
            # Limpiar la tabla antes de cargar datos
            self.tabla_novedades.delete(*self.tabla_novedades.get_children())
            if self.sheet_novedades:
                for fila in self.sheet_novedades.iter_rows(min_row=2, values_only=True):
                    # Reemplazar None por "-"
                    fila_procesada = ["-" if celda is None else celda for celda in fila]
                    self.tabla_novedades.insert("", "end", values=fila_procesada)
            else:
                print("La hoja 'NOVEDADES' está vacía o no se pudo cargar.")
        except Exception as e:
            print(f"Error al cargar los datos en el Treeview: {e}")

    # Filtrar datos según el texto ingresado
    def filtrar_datos_novedades(self, nombre_filtro,dotacion_filtro):
        try:
            # Limpiar la tabla antes de cargar datos filtrados
            self.tabla_novedades.delete(*self.tabla_novedades.get_children())
            if self.sheet_novedades:
                for fila in self.sheet_novedades.iter_rows(min_row=2, values_only=True):
                    fila_procesada = ["-" if celda is None else celda for celda in fila]
                    if dotacion_filtro == "Todas" and nombre_filtro == "Buscar por nombre":
                        self.cargar_datos_completos_novedades()
                    elif dotacion_filtro == "Todas":
                        if nombre_filtro.lower() in str(fila_procesada[3]).lower():
                            self.tabla_novedades.insert("", "end", values=fila_procesada)
                    elif nombre_filtro == "Buscar por nombre":
                        if dotacion_filtro.lower() in str(fila_procesada[5]).lower():
                            self.tabla_novedades.insert("", "end", values=fila_procesada)
                    else:
                        if nombre_filtro.lower() in str(fila_procesada[3]).lower() and dotacion_filtro.lower() in str(fila_procesada[5]).lower():
                            self.tabla_novedades.insert("", "end", values=fila_procesada)
        except Exception as e:
            print(f"Error al filtrar los datos en el Treeview: {e}")
            
    def mostrar_modal_detalle_novedades(self,novedad):
        """Mostrar el modal con el Treeview y filtro"""
        # Crear una ventana secundaria (modal)
        modal = tk.Toplevel(self.root)
        modal.title("Detalle de novedad")
        modal.geometry("560x500")
        columnas = [
            "ID", "Fecha de registro", "LEGAJO", "APELLIDOS Y NOMBRES", "ESPECIALIDAD",
            "DOTACION", "TURNOS", "FRANCO", "NOVEDAD", "Fecha de Inicio Novedad", 
            "Fecha de Fin Novedad", "REFERENCIA ESTACION", "SUPERVISOR", "Observaciones"
        ]
        
        ttk.Label(modal, text="Detalle de la novedad", font=("Helvetica", 22, "bold")).grid(row=0, column=0, columnspan=2, pady=10, padx=10, sticky="we")

        for idx, (columna, valor) in enumerate(zip(columnas, novedad), start=1):
            ttk.Label(modal, text=(columna+":").capitalize()).grid(row=idx, column=0, pady=2, padx=10, sticky="w")
            if columna == "Observaciones":
                # Usar un ScrolledText para "Observaciones"
                text_area = ScrolledText(modal, wrap=tk.WORD, height=5, width=60)
                text_area.insert("1.0", valor)
                text_area.config(state="disabled")  # Deshabilitar edición
                text_area.grid(row=idx, column=1, pady=2, padx=10, sticky="w")
            else:
                ttk.Label(modal, text=valor).grid(row=idx, column=1, pady=2, padx=10, sticky="w")
        
        # Botón de cerrar el modal
        tk.Button(modal, text="Cerrar", command=modal.destroy).grid(row=15, column=0, columnspan=2, pady=10, padx=10)        
        
    def crear_tabla_novedades(self):
        def on_focus_in(event):
            if apellido_filter.get() == "Buscar por nombre":
                apellido_filter.delete(0, tk.END)
                apellido_filter.config(fg="black")

        def on_focus_out(event):
            if apellido_filter.get() == "":
                apellido_filter.insert(0, "Buscar por nombre")
                apellido_filter.config(fg="grey")   
        def on_double_click(event):
            # Obtener el Treeview que disparó el evento
            treeview = event.widget
            # Obtener el identificador del ítem seleccionado
            selected_item = treeview.selection()

            if selected_item:  # Verificar que haya una selección
                # Obtener los valores asociados al ítem seleccionado
                item_values = treeview.item(selected_item[0], "values")
                # print(item_values)  # Imprime los valores de la fila seleccionada
                self.mostrar_modal_detalle_novedades(item_values)
            else:
                print("No se seleccionó ningún elemento.")
            
        columnas = [
            "ID", "Fecha de registro", "LEGAJO", "APELLIDOS Y NOMBRES", "ESPECIALIDAD",
            "DOTACION", "TURNOS", "FRANCO", "NOVEDAD", "Fecha de Inicio Novedad", "Fecha de Fin Novedad",
            "REFERENCIA ESTACION", "SUPERVISOR", "Observaciones"
        ]
        self.apellido_filter_var = tk.StringVar()
        self.dotacion_filter_var = tk.StringVar()
        lst_dotaciones = ["Todas","PC","LLV","TY","LP","OA","K5","RE","CÑ","AK"]
        # Detectar cambios en el input
        self.apellido_filter_var.trace_add("write", lambda *args: self.filtrar_datos_novedades(self.apellido_filter_var.get(),self.dotacion_filter_var.get()))
        self.dotacion_filter_var.trace_add("write", lambda *args: self.filtrar_datos_novedades(self.apellido_filter_var.get(),self.dotacion_filter_var.get()))
        
        # Título y botones
        ttk.Label(self.table_frame, text="Registro de novedades                  ", font=("Helvetica", 20, "bold")).grid(row=0, column=0, pady=10, padx=10, sticky="w")
        #Filter nombre o apellido
        ttk.Button(self.table_frame, text="Ver cambios de turno", command=lambda: self.toggle_view("table_cambios")).grid(row=0, column=1, pady=10, padx=10, sticky="e")
        apellido_filter = tk.Entry(self.table_frame, textvariable=self.apellido_filter_var, font=("Helvetica",10))
        apellido_filter.grid(row=0, column=2,sticky="e", pady=10, padx=10,ipady=5,ipadx=10)
        #Filter dotacion
        dotacion_filter = ttk.Combobox(self.table_frame, textvariable=self.dotacion_filter_var, values=lst_dotaciones, width=10)
        dotacion_filter.grid(row=0, column=3, sticky="e", pady=5)
        
        ttk.Button(self.table_frame, text="Nueva novedad", command=lambda: self.toggle_view("form")).grid(row=0, column=4, pady=10, padx=2, sticky="e")
        ttk.Button(self.table_frame, text="Nuevo cambio de turno", command=lambda: self.toggle_view("form_cambios")).grid(row=0, column=5, pady=10, padx=10, sticky="e")

        apellido_filter.insert(0, "Buscar por nombre")
        dotacion_filter.insert(0, "Todas")
        apellido_filter.config(fg="grey")
        apellido_filter.bind("<FocusIn>", on_focus_in)
        apellido_filter.bind("<FocusOut>", on_focus_out)
        
        # Contenedor del Treeview 
        self.tree_frame = ttk.Frame(self.table_frame, width=self.WIDTH, height=self.HEIGHT)
        # self.tree_frame = ttk.Frame(self.table_frame, width=1100, height=550)
        self.tree_frame.grid(row=1, column=0, columnspan=6, sticky="nsew")
        self.tree_frame.grid_propagate(False)

        # Configurar el grid de tree_frame para que el Treeview ocupe todo el espacio
        self.tree_frame.grid_rowconfigure(0, weight=1)  # Fila 0 del tree_frame se expande
        self.tree_frame.grid_columnconfigure(0, weight=1)  # Columna 0 del tree_frame se expande

        # Crear el Treeview
        self.tabla_novedades = ttk.Treeview(self.tree_frame, columns=columnas, show="headings", height=40)
        self.tabla_novedades.grid(row=0, column=0, sticky="nsew")

        self.tabla_novedades.bind("<Double-1>", on_double_click)
        # Establecer encabezados y anchos de columna
        anchuras = [30, 100, 60, 150, 150, 80, 60, 80, 120, 90, 120, 120, 120]
        for col, ancho in zip(columnas, anchuras):
            self.tabla_novedades.heading(col, text=col.capitalize())
            self.tabla_novedades.column(col, width=ancho)

        # Scrollbar
        scrollbar_vertical = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.tabla_novedades.yview)
        scrollbar_vertical.grid(row=0, column=1, sticky="ns")
        scrollbar_horizontal = ttk.Scrollbar(self.tree_frame, orient="horizontal", command=self.tabla_novedades.xview)
        scrollbar_horizontal.grid(row=1, column=0, sticky="ew")
        # Configurar el Treeview para usar las barras de scroll
        self.tabla_novedades.configure(yscrollcommand=scrollbar_vertical.set, xscrollcommand=scrollbar_horizontal.set)

        # Cargar datos
        self.cargar_datos_completos_novedades()        
        # Asociar el evento de doble clic

    def crear_tabla_cambios(self):
        
        columnas = [
            "ID", "Fecha de registro", "LEGAJO", "APELLIDOS Y NOMBRES", "ESPECIALIDAD", "DOTACION", 
            "TURNOS", "FRANCO", "LEGAJO2", "APELLIDOS Y NOMBRES2", "ESPECIALIDAD2", "DOTACION2", 
            "TURNOS2", "FRANCO2", "Fecha de Cambio de Turno", "REFERENCIA ESTACION", "SUPERVISOR", "Observaciones"
        ]
        # Título y botones
        ttk.Label(self.table_cambios_frame, text="Registro de cambios de turnos", font=("Helvetica", 20, "bold")).grid(row=0, column=0, pady=10, padx=10, sticky="w")
        ttk.Button(self.table_cambios_frame, text="Ver novedades", command=lambda: self.toggle_view("table")).grid(row=0, column=2, pady=10,padx=1, sticky="e")
        ttk.Button(self.table_cambios_frame, text="Nueva novedad", command=lambda: self.toggle_view("form")).grid(row=0, column=3, pady=10, padx=1, sticky="e")
        ttk.Button(self.table_cambios_frame, text="Nuevo cambio de turno", command=lambda: self.toggle_view("form_cambios")).grid(row=0, column=4, pady=10,padx=1,  sticky="e")
  
        # Contenedor del Treeview  
        self.tree_frame = ttk.Frame(self.table_cambios_frame, width=self.WIDTH, height=self.HEIGHT)
        # self.tree_frame = ttk.Frame(self.table_cambios_frame, width=1100, height=550)
        self.tree_frame.grid(row=1, column=0, columnspan=5, sticky="nsew")
        self.tree_frame.grid_propagate(False)

        # Configurar el grid del marco contenedor para que el Treeview se expanda
        self.tree_frame.grid_rowconfigure(0, weight=1)
        self.tree_frame.grid_columnconfigure(0, weight=1)

        # Crear el Treeview
        self.table_cambios = ttk.Treeview(self.tree_frame, columns=columnas, show="headings", height=38)
        self.table_cambios.grid(row=0, column=0, sticky="nsew")

        # Establecer encabezados y columnas con anchuras
        anchuras = [30, 100, 60, 150, 150, 80, 60, 80, 60, 150, 150, 80, 60, 80, 120, 120, 120, 120]
        for col, ancho in zip(columnas, anchuras):
            self.table_cambios.heading(col, text=col)
            self.table_cambios.column(col, width=ancho, anchor='center', stretch=True)

        # Scrollbar
        scrollbar_vertical = ttk.Scrollbar(self.tree_frame, orient="vertical", command=self.table_cambios.yview)
        scrollbar_vertical.grid(row=0, column=1, sticky="ns")
        scrollbar_horizontal = ttk.Scrollbar(self.tree_frame, orient="horizontal", command=self.table_cambios.xview)
        scrollbar_horizontal.grid(row=1, column=0, sticky="ew")

        # Configurar el Treeview para usar las barras de scroll
        self.table_cambios.configure(yscrollcommand=scrollbar_vertical.set, xscrollcommand=scrollbar_horizontal.set)
        # Cargar datos
        try:
            if self.sheet_cambio_turnos:
                for fila in self.sheet_cambio_turnos.iter_rows(min_row=2, values_only=True):
                    fila_procesada = ["-" if celda is None else celda for celda in fila]
                    self.table_cambios.insert("", "end", values=fila_procesada)
            else:
                print("La hoja 'NOVEDADES' está vacía o no se pudo cargar.")
        except Exception as e:
            print(f"Error al cargar los datos en el Treeview: {e}")
    
    def mostrar_formulario_novedades(self):
        
        ttk.Label(self.form_frame, text="Formulario de novedades", font=("Helvetica", 20, "bold")).grid(row=0, column=0,columnspan=3, pady=10, padx=10, sticky="nw")
        
        # Nivel 1: LEGAJO SAP
        ttk.Label(self.form_frame, text="   Legajo").grid(row=1, column=0, sticky="w")
        ttk.Label(self.form_frame, text="*", foreground="red", font=('Helvetica', 12, 'bold')).grid(row=1, column=0, sticky="w")
        self.legajo_entry = ttk.Entry(self.form_frame, textvariable=self.legajo_var, width=10)
        self.legajo_entry.grid(row=2, column=0, sticky="ew", pady=5)
        self.legajo_entry.bind("<Return>", lambda event: self.buscar_legajo())

        # Botón para buscar y auto completar
        ttk.Button(self.form_frame, text="Buscar Personal", command=self.mostrar_modal).grid(row=2, column=1, pady=10, padx=10)

        # Nivel 2: APELLIDOS Y NOMBRES, ESPECIALIDAD, DOTACION, TURNOS, FRANCO

        # Apellidos y Nombres
        ttk.Label(self.form_frame, text="  Apellido y Nombre").grid(row=3, column=0, sticky="w", padx=5)
        ttk.Label(self.form_frame, text="*", foreground="red", font=('Helvetica', 12, 'bold')).grid(row=3, column=0, sticky="w")
        self.apellidos_nombres_entry = ttk.Entry(self.form_frame, textvariable=self.apellidos_nombres_var, state='disabled', width=40)
        self.apellidos_nombres_entry.grid(row=4, column=0, sticky="w", pady=5)

        # Especialidad
        ttk.Label(self.form_frame, text="Especialidad").grid(row=3, column=1, sticky="w", padx=5)
        self.especialidad_entry = ttk.Entry(self.form_frame, textvariable=self.especialidad_var, state='disabled', width=20)
        self.especialidad_entry.grid(row=4, column=1, sticky="w", pady=5, padx=8)

        # Dotación
        ttk.Label(self.form_frame, text="Dotación").grid(row=3, column=2, sticky="w", padx=5)
        self.dotacion_entry = ttk.Entry(self.form_frame, textvariable=self.dotacion_var, state='disabled', width=20)
        self.dotacion_entry.grid(row=4, column=2, sticky="w", pady=5, padx=8)

        # Turnos
        ttk.Label(self.form_frame, text="Turno").grid(row=3, column=3, sticky="w", padx=5)
        self.turnos_entry = ttk.Entry(self.form_frame, textvariable=self.turnos_var, state='disabled', width=40)
        self.turnos_entry.grid(row=4, column=3, sticky="w", pady=5, padx=8)

        # Franco
        ttk.Label(self.form_frame, text="Franco").grid(row=3, column=4, sticky="w", padx=5)
        self.franco_entry = ttk.Entry(self.form_frame, textvariable=self.franco_var, state='disabled', width=20)
        self.franco_entry.grid(row=4, column=4, sticky="w", pady=5, padx=8)

        # Nivel 3: NOVEDAD, Fecha de Inicio Novedad
        ttk.Label(self.form_frame, text="   Tipo Novedad").grid(row=5, column=0, sticky="w")
        ttk.Label(self.form_frame, text="*", foreground="red", font=('Helvetica', 12, 'bold')).grid(row=5, column=0, sticky="w")
        self.novedad_entry = ttk.Combobox(self.form_frame, textvariable=self.novedad_var, values=self.tipo_novedades, width=38)
        self.novedad_entry.grid(row=6, column=0, columnspan=2, sticky="w", pady=5)

        ttk.Label(self.form_frame, text="   Fecha de Inicio Novedad").grid(row=5, column=3, sticky="w")
        ttk.Label(self.form_frame, text="*", foreground="red", font=('Helvetica', 12, 'bold')).grid(row=5, column=3, sticky="w")
        # Crear DateEntry con estilo ttkbootstrap
        self.fecha_inicio_novedad_entry = DateEntry(self.form_frame,dateformat='%d/%m/%Y',bootstyle="danger")
        self.fecha_inicio_novedad_entry.grid(row=6, column=3, sticky="w", pady=5)       

        ttk.Label(self.form_frame, text="Fecha de Fin Novedad").grid(row=5, column=4, sticky="w")
        # Crear DateEntry con estilo ttkbootstrap
        self.fecha_fin_novedad_entry = DateEntry(self.form_frame,dateformat='%d/%m/%Y',bootstyle="success")
        self.fecha_fin_novedad_entry.grid(row=6, column=4, sticky="w", pady=5) 
        
        # Nivel 4: REFERENCIA ESTACIÓN, SUPERVISOR
        ttk.Label(self.form_frame, text="   REFERENCIA ESTACIÓN").grid(row=7, column=0, sticky="w")
        ttk.Label(self.form_frame, text="*", foreground="red", font=('Helvetica', 12, 'bold')).grid(row=7, column=0, sticky="w")
        self.referencia_estacion_entry = ttk.Entry(self.form_frame, textvariable=self.referencia_estacion_var, width=40)
        self.referencia_estacion_entry.grid(row=8, column=0, columnspan=2, sticky="w", pady=5)

        ttk.Label(self.form_frame, text="   SUPERVISOR").grid(row=7, column=3, sticky="w")
        ttk.Label(self.form_frame, text="*", foreground="red", font=('Helvetica', 12, 'bold')).grid(row=7, column=3, sticky="w")
        self.supervisor_entry = ttk.Entry(self.form_frame, textvariable=self.supervisor_var, width=40)
        self.supervisor_entry.grid(row=8, column=3, columnspan=2, sticky="w", pady=5)

        # Nivel 5: Observaciones
        ttk.Label(self.form_frame, text="Observaciones").grid(row=9, column=0, sticky="w")
        self.observaciones_text = scrolledtext.ScrolledText(self.form_frame, wrap=tk.WORD, height=12,width=180)
        self.observaciones_text.grid(row=10, column=0, columnspan=7, pady=5, sticky="w")

        # Botón de Guardar Datos en el formulario
        ttk.Button(self.form_frame, text="Guardar Novedad", command=self.guardar_datos_novedades).grid(row=11, column=3, columnspan=2, pady=10)

        # Botón para cerrar el formulario y regresar a la tabla
        ttk.Button(self.form_frame, text="Cerrar", command=lambda: (self.limpiar_formulario_novedades(), self.toggle_view())).grid(row=11, column=4, columnspan=2, pady=10)

        
    def mostrar_formulario_cambios(self):
        ttk.Label(self.form_cambios_frame, text="Formulario de cambios de turnos", font=("Helvetica", 20, "bold")).grid(row=0, column=0,columnspan=3, pady=10, padx=10, sticky="nw")

        # Nivel 1: LEGAJO SAP
        ttk.Label(self.form_cambios_frame, text="   Legajo").grid(row=1, column=0, sticky="w")# Ejemplo de cómo agregar un asterisco rojo a la etiqueta
        ttk.Label(self.form_cambios_frame, text="*", foreground="red", font=('Helvetica', 12, 'bold')).grid(row=1, column=0, sticky="w")
        self.legajo_entry = ttk.Entry(self.form_cambios_frame, textvariable=self.legajo_var, width=10)
        self.legajo_entry.grid(row=2, column=0, sticky="w", pady=5)
        self.legajo_entry.bind("<Return>", lambda event: self.buscar_legajo())

        # Botón para buscar y auto completar
        ttk.Button(self.form_cambios_frame, text="Buscar Personal", command=lambda: self.mostrar_modal(1)).grid(row=2, column=1, pady=10, padx=10)

        # Nivel 2: APELLIDOS Y NOMBRES, ESPECIALIDAD, DOTACION, TURNOS, FRANCO
        # Apellidos y Nombres
        ttk.Label(self.form_cambios_frame, text="  Apellido y Nombre").grid(row=3, column=0, sticky="w", padx=5)
        ttk.Label(self.form_cambios_frame, text="*", foreground="red", font=('Helvetica', 12, 'bold')).grid(row=3, column=0, sticky="w")
        self.apellidos_nombres_entry = ttk.Entry(self.form_cambios_frame, textvariable=self.apellidos_nombres_var, state='disabled', width=40)
        self.apellidos_nombres_entry.grid(row=4, column=0,columnspan = 2, sticky="w", pady=5)

        # Especialidad
        ttk.Label(self.form_cambios_frame, text="Especialidad").grid(row=3, column=2, sticky="w", padx=5)
        self.especialidad_entry = ttk.Entry(self.form_cambios_frame, textvariable=self.especialidad_var, state='disabled', width=20)
        self.especialidad_entry.grid(row=4, column=2, sticky="w", pady=5, padx=8)

        # Dotación
        ttk.Label(self.form_cambios_frame, text="Dotación").grid(row=3, column=3, sticky="w", padx=5)
        self.dotacion_entry = ttk.Entry(self.form_cambios_frame, textvariable=self.dotacion_var, state='disabled', width=20)
        self.dotacion_entry.grid(row=4, column=3, sticky="w", pady=5, padx=8)

        # Turnos
        ttk.Label(self.form_cambios_frame, text="Turno").grid(row=3, column=4, sticky="w", padx=5)
        self.turnos_entry = ttk.Entry(self.form_cambios_frame, textvariable=self.turnos_var, state='disabled', width=40)
        self.turnos_entry.grid(row=4, column=4, columnspan= 2 , sticky="w", pady=5, padx=8)

        # Franco
        ttk.Label(self.form_cambios_frame, text="Franco").grid(row=3, column=6, sticky="w", padx=5)
        self.franco_entry = ttk.Entry(self.form_cambios_frame, textvariable=self.franco_var, state='disabled', width=20)
        self.franco_entry.grid(row=4, column=6, sticky="w", pady=5, padx=8)
        #--------------------------------------------------------------------------------
        # Nivel 3: LEGAJO 2
        ttk.Label(self.form_cambios_frame, text="   Legajo 2").grid(row=5, column=0, sticky="w")# Ejemplo de cómo agregar un asterisco rojo a la etiqueta
        ttk.Label(self.form_cambios_frame, text="*", foreground="red", font=('Helvetica', 12, 'bold')).grid(row=5, column=0, sticky="w")
        self.legajo_2_entry = ttk.Entry(self.form_cambios_frame, textvariable=self.legajo_2_var, width=10)
        self.legajo_2_entry.grid(row=6, column=0, sticky="w", pady=5)
        self.legajo_2_entry.bind("<Return>", lambda event: self.buscar_legajo(2))

        # Botón para buscar y auto completar
        ttk.Button(self.form_cambios_frame, text="Buscar Personal", command=lambda: self.mostrar_modal(2)).grid(row=6, column=1, pady=10, padx=10)

        # Nivel 4: APELLIDOS Y NOMBRES, ESPECIALIDAD, DOTACION, TURNOS, FRANCO
        # Apellidos y Nombres2
        ttk.Label(self.form_cambios_frame, text="  Apellido y Nombre 2").grid(row=7, column=0, sticky="w", padx=5)
        ttk.Label(self.form_cambios_frame, text="*", foreground="red", font=('Helvetica', 12, 'bold')).grid(row=7, column=0, sticky="w")
        self.apellidos_nombres_entry_2 = ttk.Entry(self.form_cambios_frame, textvariable=self.apellidos_nombres_2_var, state='disabled', width=40)
        self.apellidos_nombres_entry_2.grid(row=8, column=0,columnspan = 2, sticky="w", pady=5)

        # Especialidad2
        ttk.Label(self.form_cambios_frame, text="Especialidad 2").grid(row=7, column=2, sticky="w", padx=5)
        self.especialidad_entry_2 = ttk.Entry(self.form_cambios_frame, textvariable=self.especialidad_2_var, state='disabled', width=20)
        self.especialidad_entry_2.grid(row=8, column=2, sticky="w", pady=5, padx=8)

        # Dotación2
        ttk.Label(self.form_cambios_frame, text="Dotación 2").grid(row=7, column=3, sticky="w", padx=5)
        self.dotacion_entry_2 = ttk.Entry(self.form_cambios_frame, textvariable=self.dotacion_2_var, state='disabled', width=20)
        self.dotacion_entry_2.grid(row=8, column=3, sticky="w", pady=5, padx=8)

        # Turnos2
        ttk.Label(self.form_cambios_frame, text="Turno 2").grid(row=7, column=4, sticky="w", padx=5)
        self.turnos_entry_2 = ttk.Entry(self.form_cambios_frame, textvariable=self.turnos_2_var, state='disabled', width=40)
        self.turnos_entry_2.grid(row=8, column=4, columnspan= 2 , sticky="w", pady=5, padx=8)

        # Franco2
        ttk.Label(self.form_cambios_frame, text="Franco 2").grid(row=7, column=6, sticky="w", padx=5)
        self.franco_entry_2 = ttk.Entry(self.form_cambios_frame, textvariable=self.franco_2_var, state='disabled', width=20)
        self.franco_entry_2.grid(row=8, column=6, sticky="w", pady=5, padx=8)
        
        # Nivel 5: Fecha de cambio de turno
        ttk.Label(self.form_cambios_frame, text="   Fecha de cambio de turno").grid(row=9, column=0, sticky="w")
        ttk.Label(self.form_cambios_frame, text="*", foreground="red", font=('Helvetica', 12, 'bold')).grid(row=9, column=0, sticky="w")
        # Crear DateEntry con estilo ttkbootstrap
        self.fecha_cambio_turno_entry = DateEntry(self.form_cambios_frame,dateformat='%d/%m/%Y',startdate=datetime.today()+timedelta(days=+1))
        self.fecha_cambio_turno_entry.grid(row=10, column=0, columnspan=2, sticky="w", pady=5)
        
        # Nivel 6: REFERENCIA ESTACIÓN, SUPERVISOR
        ttk.Label(self.form_cambios_frame, text="   REFERENCIA ESTACIÓN").grid(row=11, column=0, sticky="w")
        ttk.Label(self.form_cambios_frame, text="*", foreground="red", font=('Helvetica', 12, 'bold')).grid(row=11, column=0, sticky="w")
        self.referencia_estacion_entry = ttk.Entry(self.form_cambios_frame, textvariable=self.referencia_estacion_var, width=40)
        self.referencia_estacion_entry.grid(row=12, column=0, columnspan=2, sticky="w", pady=5)

        ttk.Label(self.form_cambios_frame, text="   SUPERVISOR").grid(row=11, column=3, sticky="w")
        ttk.Label(self.form_cambios_frame, text="*", foreground="red", font=('Helvetica', 12, 'bold')).grid(row=11, column=3, sticky="w")
        self.supervisor_entry = ttk.Entry(self.form_cambios_frame, textvariable=self.supervisor_var, width=40)
        self.supervisor_entry.grid(row=12, column=3, columnspan=2, sticky="w", pady=5)

        # Nivel 7: Observaciones
        ttk.Label(self.form_cambios_frame, text="Observaciones").grid(row=13, column=0, sticky="w")
        self.observaciones_text = scrolledtext.ScrolledText(self.form_cambios_frame, wrap=tk.WORD, height=1,width=170)
        self.observaciones_text.grid(row=14, column=0, columnspan=7, pady=5, sticky="w")

        # Botón de Guardar Datos en el formulario
        ttk.Button(self.form_cambios_frame, text="Guardar cambio de turno", command=self.guardar_datos_cambios).grid(row=15, column=2, columnspan=2, pady=10)

        # Botón para cerrar el formulario y regresar a la tabla
        ttk.Button(self.form_cambios_frame, text="Cerrar", command=lambda: (self.limpiar_formulario_novedades(), self.toggle_view("table_cambios"))).grid(row=15, column=3, columnspan=2, pady=10)
            
    def mostrar_modal(self,boton=1):
        """Mostrar el modal con el Treeview y filtro"""
        # Crear una ventana secundaria (modal)
        modal = tk.Toplevel(self.root)
        modal.title("Seleccionar Legajo")
        modal.geometry("1250x500")
        
        # Crear un campo de entrada para filtrar por apellido
        tk.Label(modal, text="Filtrar por apellido o nombre:").grid(row=0, column=0, padx=10, pady=5)
        # Crear el Entry con ttkbootstrap
        # apellido_filter = Entry(modal)
        apellido_filter_var = tk.StringVar()
        apellido_filter = tk.Entry(modal, textvariable=apellido_filter_var)
        apellido_filter.grid(row=0, column=1, padx=10, pady=5,ipady=5,ipadx=50)
        
        # Crear el Treeview para mostrar la tabla
        tree = ttk.Treeview(modal, columns=("legajo", "apellido_nombre", "especialidad", "dotacion", "turnos", "franco"), show="headings",height=25)
        tree.heading("legajo", text="LEGAJO SAP")
        tree.heading("apellido_nombre", text="APELLIDOS Y NOMBRES")
        tree.heading("especialidad", text="ESPECIALIDAD")
        tree.heading("dotacion", text="DOTACION")
        tree.heading("turnos", text="TURNOS")
        tree.heading("franco", text="FRANCO")
        tree.grid(row=1, column=0, columnspan=6, padx=20, pady=10)

        # Función para llenar la tabla con los datos de la hoja 'BASE'
        def cargar_tabla():
            for row in tree.get_children():
                tree.delete(row)
            
            apellido_filter = apellido_filter_var.get().lower()  # Obtener el filtro (en minúsculas)
            
            # Cargar las filas del Excel y mostrar en el Treeview
            for row in self.sheet_base.iter_rows(min_row=2, values_only=True):
                legajo = row[0]
                apellidos_nombres = row[1]
                turnos = row[2]
                franco = row[3]
                especialidad = row[4]
                dotacion = row[5]

                if apellido_filter in apellidos_nombres.lower():  # Comparar sin tener en cuenta mayúsculas
                    tree.insert("", "end", values=(legajo, apellidos_nombres, especialidad, dotacion, turnos, franco))
        # Llamar a la función para cargar la tabla al inicio
        cargar_tabla()
        # Asociar el evento de filtrar al campo de filtro por apellido
        apellido_filter_var.trace_add("write", lambda *args: cargar_tabla())

        def on_double_click(event):
            # Función para asignar el legajo al campo de entrada y cerrar el modal
            selected_item = tree.selection()[0]  # Obtener el item seleccionado
            legajo_sap = tree.item(selected_item, "values")[0]  # Obtener el legajo
            apellido_nombre = tree.item(selected_item, "values")[1]  
            especialidad = tree.item(selected_item, "values")[2]
            dotacion = tree.item(selected_item, "values")[3]
            turnos = tree.item(selected_item, "values")[4]
            franco = tree.item(selected_item, "values")[5]
            if boton == 1:
                self.legajo_var.set(legajo_sap)  # Asignar al campo "LEGAJO SAP"
                self.apellidos_nombres_var.set(apellido_nombre)  
                self.especialidad_var.set(especialidad)  
                self.dotacion_var.set(dotacion)  
                self.turnos_var.set(turnos)  
                self.franco_var.set(franco)  
            elif boton == 2:
                self.legajo_2_var.set(legajo_sap)  # Asignar al campo "LEGAJO SAP"
                self.apellidos_nombres_2_var.set(apellido_nombre)  
                self.especialidad_2_var.set(especialidad)  
                self.dotacion_2_var.set(dotacion)  
                self.turnos_2_var.set(turnos)  
                self.franco_2_var.set(franco) 
                
            modal.destroy()  # Cerrar el modal

        # Asociar el evento de doble clic
        tree.bind("<Double-1>", on_double_click)

        # Botón de cerrar el modal
        tk.Button(modal, text="Cerrar", command=modal.destroy).grid(row=2, column=0, columnspan=2, pady=10)

    def buscar_legajo(self,campo=1):
        """Buscar el legajo SAP en la hoja 'BASE' y auto completar los datos"""
        try:
            # Convertir el legajo ingresado a número (int)
            if campo == 1:
                legajo = int(self.legajo_var.get().strip())  # Eliminar espacios y convertir a número
            elif campo == 2:
                legajo = int(self.legajo_2_var.get().strip())  # Eliminar espacios y convertir a número
                
            print(f"Buscando legajo SAP: {legajo}")

            # Buscar en la hoja BASE (salta el encabezado en la primera fila)
            for row in self.sheet_base.iter_rows(min_row=2, values_only=True):
                # Convertir el legajo de la hoja a número (int)
                legajo_sap = int(row[0])  
                if legajo_sap == legajo:
                    # Si el legajo SAP coincide, auto completar los campos
                    if campo == 1:
                        self.apellidos_nombres_var.set(row[1])  # Nombre en la segunda columna
                        self.turnos_var.set(row[2])  # Turnos
                        self.franco_var.set(row[3])  # Franco
                        self.especialidad_var.set(row[4])  # Especialidad
                        self.dotacion_var.set(row[5])  # Dotación
                    elif campo == 2:
                        self.apellidos_nombres_2_var.set(row[1])  # Nombre en la segunda columna
                        self.turnos_2_var.set(row[2])  # Turnos
                        self.franco_2_var.set(row[3])  # Franco
                        self.especialidad_2_var.set(row[4])  # Especialidad
                        self.dotacion_2_var.set(row[5])  # Dotación
                    print(f"Legajo encontrado: {legajo_sap}")
                    # self.deshabilitar_campos()  # Deshabilitar los campos después de la búsqueda
                    break  # Salir del bucle una vez encontrado el legajo
            else:
                messagebox.showinfo("Legajo No Encontrado", f"El legajo {legajo} no fue encontrado.")
                print(f"Legajo SAP {legajo} no encontrado.")

        except ValueError:
            # Si el legajo ingresado no es un número, mostrar un mensaje de error
            messagebox.showerror("Error de entrada", "Por favor, ingrese un número de legajo válido.")

    def guardar_datos_novedades(self):
        """Guardar los datos del formulario en el archivo Excel"""
        # self.cargar_excel()
        self.fecha_inicio_novedad_var.set(self.fecha_inicio_novedad_entry.entry.get())
        self.fecha_fin_novedad_var.set(self.fecha_fin_novedad_entry.entry.get())
        self.observaciones_var.set(self.observaciones_text.get("1.0", "end-1c"))

        # Obtener el último ID válido en la columna 1, omitiendo el encabezado
        last_id = max(
            (row[0] for row in self.sheet_novedades.iter_rows(min_row=2, max_col=1, values_only=True) if isinstance(row[0], int)),
            default=0
        )
        # Incrementar el ID en 1
        new_id = int(last_id) + 1 if last_id else 1
        
        # Obtener la fecha y hora actual en el formato especificado
        current_datetime = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        if self.validar_campos_requeridos_novedades():
            # Aquí guarda los datos si todo está bien
            # Puedes agregar el código para guardar en el archivo de Excel o donde necesites.
            # Añadir a la hoja de Excel y guardar
            try:
                # Añadir los nuevos datos a la hoja
                self.sheet_novedades.insert_rows(2)
                self.sheet_novedades['A2'] = new_id
                self.sheet_novedades['B2'] = current_datetime
                self.sheet_novedades['C2'] = self.legajo_var.get()
                self.sheet_novedades['D2'] = self.apellidos_nombres_var.get()
                self.sheet_novedades['E2'] = self.especialidad_var.get()
                self.sheet_novedades['F2'] = self.dotacion_var.get()
                self.sheet_novedades['G2'] = self.turnos_var.get()
                self.sheet_novedades['H2'] = self.franco_var.get()
                self.sheet_novedades['I2'] = self.novedad_var.get()
                self.sheet_novedades['J2'] = self.fecha_inicio_novedad_var.get()
                self.sheet_novedades['K2'] = self.fecha_fin_novedad_var.get()
                self.sheet_novedades['L2'] = self.referencia_estacion_var.get()
                self.sheet_novedades['M2'] = self.supervisor_var.get()
                self.sheet_novedades['N2'] = self.observaciones_var.get()
                
                # Guardar el archivo Excel
                self.wb.save(self.excel_file)

                # Mensaje de confirmación
                messagebox.showinfo("Guardado", "Los datos han sido guardados correctamente.")

                # Limpiar el formulario y alternar a la vista de tabla
                self.limpiar_formulario_novedades()
                self.toggle_view()
                print("Datos guardados correctamente.")
            except PermissionError:
                messagebox.showerror("Error", "No se pudo guardar el archivo porque está abierto en otro programa. Por favor, cierre el archivo y vuelva a intentarlo.")
            except Exception as e:
                # Mensaje de error si ocurre un problema
                messagebox.showerror("Error", f"No se pudieron guardar los datos: {str(e)} Por favor, intente de nuevo y si el problema persiste avise al administrador")
                print(f"Error al guardar los datos: {e}")
            
        else:
            print("Algunos campos son obligatorios y están vacíos.")
            messagebox.showinfo("Faltan Datos requeridos", f"Algunos campos son obligatorios y están vacíos.")
    
    def guardar_datos_cambios(self):
        """Guardar los datos del formulario en el archivo Excel"""
        # self.cargar_excel()
        self.fecha_cambio_turno_var.set(self.fecha_cambio_turno_entry.entry.get())
        self.observaciones_var.set(self.observaciones_text.get("1.0", "end-1c"))

        # Obtener el último ID válido en la columna 1, omitiendo el encabezado    
        last_id = max(
            (row[0] for row in self.sheet_cambio_turnos.iter_rows(min_row=2, max_col=1, values_only=True) if isinstance(row[0], int)),
            default=0
        )

        # Incrementar el ID en 1
        new_id = int(last_id) + 1 if last_id else 1

        # Obtener la fecha y hora actual en el formato especificado
        current_datetime = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        if self.validar_campos_requeridos_cambios():
            # Aquí guarda los datos si todo está bien
            # Puedes agregar el código para guardar en el archivo de Excel o donde necesites.
            # Añadir a la hoja de Excel y guardar
            try:
                # Añadir los nuevos datos a la hoja
                self.sheet_cambio_turnos.insert_rows(2)
                self.sheet_cambio_turnos['A2'] = new_id
                self.sheet_cambio_turnos['B2'] = current_datetime
                self.sheet_cambio_turnos['C2'] = self.legajo_var.get()
                self.sheet_cambio_turnos['D2'] = self.apellidos_nombres_var.get()
                self.sheet_cambio_turnos['E2'] = self.especialidad_var.get()
                self.sheet_cambio_turnos['F2'] = self.dotacion_var.get()
                self.sheet_cambio_turnos['G2'] = self.turnos_var.get()
                self.sheet_cambio_turnos['H2'] = self.franco_var.get()
                self.sheet_cambio_turnos['I2'] = self.legajo_2_var.get()
                self.sheet_cambio_turnos['J2'] = self.apellidos_nombres_2_var.get()
                self.sheet_cambio_turnos['K2'] = self.especialidad_2_var.get()
                self.sheet_cambio_turnos['L2'] = self.dotacion_2_var.get()
                self.sheet_cambio_turnos['M2'] = self.turnos_2_var.get()
                self.sheet_cambio_turnos['N2'] = self.franco_2_var.get()
                self.sheet_cambio_turnos['O2'] = self.fecha_cambio_turno_var.get()
                self.sheet_cambio_turnos['P2'] = self.referencia_estacion_var.get()
                self.sheet_cambio_turnos['Q2'] = self.supervisor_var.get()
                self.sheet_cambio_turnos['R2'] = self.observaciones_var.get()

                # Guardar el archivo Excel
                self.wb.save(self.excel_file)

                # Mensaje de confirmación
                messagebox.showinfo("Guardado", "Los datos han sido guardados correctamente.")

                # Limpiar el formulario y alternar a la vista de tabla
                self.limpiar_formulario_cambios()
                self.toggle_view("table_cambios")
                print("Datos guardados correctamente.")
            except PermissionError:
                messagebox.showerror("Error", "No se pudo guardar el archivo porque está abierto en otro programa. Por favor, cierre el archivo y vuelva a intentarlo.")
            except Exception as e:
                # Mensaje de error si ocurre un problema

                messagebox.showerror("Error", f"No se pudieron guardar los datos: {str(e)} Por favor, intente de nuevo y si el problema persiste avise al administrador")
                print(f"Error al guardar los datos: {e}")
        else:
            print("Algunos campos son obligatorios y están vacíos.")
            messagebox.showinfo("Faltan Datos requeridos", f"Algunos campos son obligatorios y están vacíos.")
    def limpiar_formulario_novedades(self):
        self.legajo_var.set('')
        self.apellidos_nombres_var.set('')
        self.especialidad_var.set('')
        self.dotacion_var.set('')
        self.turnos_var.set('')
        self.franco_var.set('')
        self.novedad_var.set('')        
        self.fecha_inicio_novedad_var.set('')
        self.fecha_fin_novedad_var.set('')
        self.referencia_estacion_var.set('')
        self.supervisor_var.set('')
        self.observaciones_var.set('')
        self.observaciones_text.delete("1.0", "end")

    
    def limpiar_formulario_cambios(self):
        self.legajo_var.set('')
        self.apellidos_nombres_var.set('')
        self.especialidad_var.set('')
        self.dotacion_var.set('')
        self.turnos_var.set('')
        self.franco_var.set('')
        self.legajo_2_var.set('')
        self.apellidos_nombres_2_var.set('')
        self.especialidad_2_var.set('')
        self.dotacion_2_var.set('')
        self.turnos_2_var.set('')
        self.franco_2_var.set('')
        self.fecha_cambio_turno_var.set('')
        self.referencia_estacion_var.set('')
        self.supervisor_var.set('')
        self.observaciones_var.set('') 
        self.observaciones_text.delete("1.0", "end")
   
    def validar_campos_requeridos_novedades(self):
            # Verificar si los campos requeridos están vacíos
            if not self.legajo_var.get():
                self.mostrar_error_novedades("El campo 'Legajo' es obligatorio.")
                return False
            if not self.apellidos_nombres_var.get():
                self.mostrar_error_novedades("El campo 'Apellido y Nombre' es obligatorio.")
                return False
            if not self.novedad_var.get():
                self.mostrar_error_novedades("El campo 'Tipo de Novedad' es obligatorio.")
                return False
            if not self.fecha_inicio_novedad_var.get():
                self.mostrar_error_novedades("El campo 'Fecha de inicio novedad' es obligatorio.")
                return False
            if not self.referencia_estacion_var.get():
                self.mostrar_error_novedades("El campo 'Referencia Estacion' es obligatorio.")
                return False
            if not self.supervisor_var.get():
                self.mostrar_error_novedades("El campo 'Supervisor' es obligatorio.")
                return False
            # Verifica otros campos de la misma forma...

            return True
        
    def validar_campos_requeridos_cambios(self):
            # Verificar si los campos requeridos están vacíos
            if not self.legajo_var.get():
                self.mostrar_error_cambios("El campo 'Legajo' es obligatorio.")
                return False
            if not self.apellidos_nombres_var.get():
                self.mostrar_error_cambios("El campo 'Apellido y Nombre' es obligatorio.")
                return False
            if not self.legajo_2_var.get():
                self.mostrar_error_cambios("El campo 'Legajo' es obligatorio.")
                return False
            if not self.apellidos_nombres_2_var.get():
                self.mostrar_error_cambios("El campo 'Apellido y Nombre' es obligatorio.")
                return False
            if not self.fecha_cambio_turno_var.get():
                self.mostrar_error_cambios("El campo 'Fecha de cambio de turno' es obligatorio.")
                return False
            if not self.referencia_estacion_var.get():
                self.mostrar_error_cambios("El campo 'Referencia Estacion' es obligatorio.")
                return False
            if not self.supervisor_var.get():
                self.mostrar_error_cambios("El campo 'Supervisor' es obligatorio.")
                return False
            # Verifica otros campos de la misma forma...

            return True
    def mostrar_error_novedades(self, mensaje):
        """Mostrar mensaje de error."""
        error_label = ttk.Label(self.form_frame, text=mensaje, foreground="red")
        error_label.grid(row=11, column=0, columnspan=7, pady=5, sticky="w")
    
    def mostrar_error_cambios(self, mensaje):
        """Mostrar mensaje de error."""
        error_label = ttk.Label(self.form_cambios_frame, text=mensaje, foreground="red")
        error_label.grid(row=15, column=0, columnspan=7, pady=5, sticky="w")

# Crear la ventana principal de Tkinter
if __name__ == "__main__":
    root = tk.Tk()
    app = FormularioExcelApp(root)
    # Crear un hilo para ejecutar la tarea periódica sin bloquear la interfaz
    hilo_periodico = threading.Thread(target=app.ejecutar_periodicamente)
    hilo_periodico.daemon = True  # Esto permite que el hilo se cierre cuando la aplicación principal termine
    hilo_periodico.start()

    # Iniciar el bucle principal de Tkinter
    root.mainloop()