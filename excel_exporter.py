"""Exportación de las tablas operativas SQLite a un libro Excel."""

from pathlib import Path
import os
import tempfile

import openpyxl
from openpyxl.utils import get_column_letter


def _write_sheet(workbook, title, headers, rows):
    worksheet = workbook.create_sheet(title=title)
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row[key] for key in row.keys()])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, header in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = max(12, min(35, len(header) + 2))


def _clave_fecha(value):
    """Normaliza una fecha ISO (YYYY-MM-DD) o DD/MM/YYYY a YYYYMMDD comparable.

    Devuelve "" si el valor no puede normalizarse.
    """
    value = (value or "").strip()
    if len(value) < 10:
        return ""
    if value[4] == "-" and value[7] == "-":
        return value[0:4] + value[5:7] + value[8:10]
    if value[2] == "/" and value[5] == "/":
        return value[6:10] + value[3:5] + value[0:2]
    return ""


def export_database(store, output_path, fecha_desde=None, fecha_hasta=None,
                    id_desde=None, id_hasta=None, tipo_novedad=None, tables=None):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    selected_tables = set(tables or ("BASE", "TipoNovedad", "NOVEDADES", "Cambio de Turnos"))
    with store.read_connection() as connection:
        base = connection.execute(
            """SELECT legajo, apellidos_nombres, especialidad, dotacion, turnos, franco
               FROM empleados WHERE activo=1 ORDER BY apellidos_nombres"""
        ).fetchall() if "BASE" in selected_tables else []
        tipos = connection.execute(
            "SELECT nombre FROM tipos_novedad WHERE activo=1 ORDER BY nombre"
        ).fetchall() if "TipoNovedad" in selected_tables else []
        common_conditions = []
        common_params = []
        if id_desde is not None:
            common_conditions.append("id >= ?")
            common_params.append(id_desde)
        if id_hasta is not None:
            common_conditions.append("id <= ?")
            common_params.append(id_hasta)
        common_conditions.append("activo = 1")
        where_common = " AND ".join(common_conditions) or "1=1"
        novedades_conditions = list(common_conditions)
        novedades_params = list(common_params)
        if tipo_novedad:
            novedades_conditions.append("novedad = ?")
            novedades_params.append(tipo_novedad)
        novedades = connection.execute(
            """SELECT id, registrado_en, legajo, apellidos_nombres, especialidad, dotacion,
                      turnos, franco, novedad, fecha_inicio, fecha_fin, referencia_estacion,
                      supervisor, observaciones, usuario_windows
               FROM novedades WHERE """ + " AND ".join(novedades_conditions or ["1=1"]) + " ORDER BY id DESC",
            novedades_params,
        ).fetchall() if "NOVEDADES" in selected_tables else []
        cambios = connection.execute(
            """SELECT id, registrado_en, legajo_1, apellidos_nombres_1, especialidad_1,
                      dotacion_1, turnos_1, franco_1, legajo_2, apellidos_nombres_2,
                      especialidad_2, dotacion_2, turnos_2, franco_2, fecha_cambio,
                      referencia_estacion, supervisor, observaciones, usuario_windows
               FROM cambios_turno WHERE """ + where_common + " ORDER BY id DESC""",
            common_params,
        ).fetchall() if "Cambio de Turnos" in selected_tables else []
        personal = [
            {"NOMBRE": fila["nombre"], "ACTIVO": "Sí" if fila["activo"] else "No"}
            for fila in connection.execute(
                "SELECT nombre, activo FROM personal_estacion ORDER BY nombre"
            ).fetchall()
        ] if "PersonalEstacion" in selected_tables else []

    desde_clave = _clave_fecha(fecha_desde)
    hasta_clave = _clave_fecha(fecha_hasta)

    def _en_rango(registrado_en):
        if not desde_clave and not hasta_clave:
            return True
        clave = _clave_fecha(registrado_en)
        if not clave:
            return False
        if desde_clave and clave < desde_clave:
            return False
        if hasta_clave and clave > hasta_clave:
            return False
        return True

    novedades = [fila for fila in novedades if _en_rango(fila["registrado_en"])]
    cambios = [fila for fila in cambios if _en_rango(fila["registrado_en"])]

    workbook = openpyxl.Workbook()
    del workbook[workbook.sheetnames[0]]
    if "BASE" in selected_tables:
        _write_sheet(workbook, "BASE", ["LEGAJO", "APELLIDOS Y NOMBRES", "ESPECIALIDAD", "DOTACION", "TURNOS", "FRANCO"], base)
    if "TipoNovedad" in selected_tables:
        _write_sheet(workbook, "TipoNovedad", ["NOVEDAD"], tipos)
    if "NOVEDADES" in selected_tables:
        _write_sheet(workbook, "NOVEDADES", [
            "ID", "Fecha y hora", "LEGAJO", "APELLIDOS Y NOMBRES", "ESPECIALIDAD", "DOTACION",
            "TURNOS", "FRANCO", "NOVEDAD", "Fecha de Inicio Novedad", "Fecha de Fin Novedad",
            "REFERENCIA ESTACIÓN", "SUPERVISOR", "Observaciones", "USUARIO WINDOWS"
        ], novedades)
    if "Cambio de Turnos" in selected_tables:
        _write_sheet(workbook, "Cambio de Turnos", [
            "ID", "Fecha y hora", "LEGAJO", "APELLIDOS Y NOMBRES", "ESPECIALIDAD", "DOTACION",
            "TURNOS", "FRANCO", "LEGAJO2", "APELLIDOS Y NOMBRES2", "ESPECIALIDAD2", "DOTACION2",
            "TURNOS2", "FRANCO2", "Fecha de Cambio de Turno", "REFERENCIA ESTACIÓN", "SUPERVISOR",
            "Observaciones", "USUARIO WINDOWS"
        ], cambios)
    if "PersonalEstacion" in selected_tables:
        _write_sheet(workbook, "PersonalEstacion", ["NOMBRE", "ACTIVO"], personal)

    fd, temporary_path = tempfile.mkstemp(prefix="registro_export_", suffix=".xlsx", dir=output_path.parent)
    os.close(fd)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
    finally:
        if os.path.exists(temporary_path):
            os.remove(temporary_path)
    return str(output_path)


def export_auditoria(store, output_path, text_filter=""):
    """Exporta la auditoría a un Excel independiente, aplicando el filtro actual."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    like = f"%{(text_filter or '').strip()}%"
    with store.read_connection() as connection:
        rows = connection.execute(
            """SELECT a.id, a.creado_en, COALESCE(u.username, a.usuario_windows, '') AS usuario,
                      a.accion, a.entidad, a.entidad_id, a.datos_anteriores, a.datos_nuevos
               FROM auditoria a LEFT JOIN usuarios u ON u.id=a.usuario_id
               WHERE ?='' OR COALESCE(u.username, a.usuario_windows, '') LIKE ?
                     OR a.accion LIKE ? OR a.entidad LIKE ?
               ORDER BY a.id DESC""",
            ((text_filter or '').strip(), like, like, like),
        ).fetchall()
    workbook = openpyxl.Workbook()
    del workbook[workbook.sheetnames[0]]
    _write_sheet(
        workbook,
        "Auditoria",
        ["ID", "Fecha", "Usuario", "Acción", "Entidad", "ID registro", "Datos anteriores", "Datos nuevos"],
        rows,
    )
    fd, temporary_path = tempfile.mkstemp(prefix="registro_auditoria_", suffix=".xlsx", dir=output_path.parent)
    os.close(fd)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
    finally:
        if os.path.exists(temporary_path):
            os.remove(temporary_path)
    return str(output_path)
