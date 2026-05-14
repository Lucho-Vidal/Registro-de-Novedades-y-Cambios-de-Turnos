import os
import getpass
import openpyxl


def get_workbook_mtime(file_path):
    try:
        return os.path.getmtime(file_path)
    except OSError:
        return None


def load_workbook_if_needed(file_path, last_mtime=None, only_if_changed=False):
    mtime_actual = get_workbook_mtime(file_path)
    if only_if_changed and last_mtime is not None and mtime_actual == last_mtime:
        return None, mtime_actual, False

    wb = openpyxl.load_workbook(file_path)
    return wb, mtime_actual, True


def build_base_cache(sheet_base):
    base_rows = []
    base_index = {}
    for row in sheet_base.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        base_rows.append(row)
        try:
            legajo = int(row[0])
            base_index[legajo] = row
        except (TypeError, ValueError):
            continue
    return base_rows, base_index


def get_windows_user():
    try:
        return getpass.getuser()
    except Exception:
        return os.environ.get("USERNAME", "")


def ensure_user_column(sheet, title="USUARIO WINDOWS"):
    encabezados = [cell.value for cell in sheet[1]]
    if title in encabezados:
        return encabezados.index(title) + 1
    columna_nueva = sheet.max_column + 1
    sheet.cell(row=1, column=columna_nueva, value=title)
    return columna_nueva


def get_last_id(sheet):
    return max(
        (row[0] for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True) if isinstance(row[0], int)),
        default=0,
    )


def create_default_workbook(file_path, sheet_base, sheet_novedades, sheet_tipo_novedad, sheet_cambio_turnos, col_usuario):
    wb = openpyxl.Workbook()

    ws_base = wb.create_sheet(title=sheet_base)
    ws_base.append(["LEGAJO", "APELLIDOS Y NOMBRES", "ESPECIALIDAD", "DOTACION", "TURNOS", "FRANCO"])

    ws_nov = wb.create_sheet(title=sheet_novedades)
    ws_nov.append([
        "ID", "Fecha y hora", "LEGAJO ", "APELLIDOS Y NOMBRES", "ESPECIALIDAD",
        "DOTACION", "TURNOS", "FRANCO", "NOVEDAD", "Fecha de Inicio Novedad",
        "Fecha de Fin Novedad", "REFERENCIA ESTACIÓN", "SUPERVISOR", "Observaciones", col_usuario,
    ])

    ws_tipo = wb.create_sheet(title=sheet_tipo_novedad)
    ws_tipo.append(["Enfermo"])

    ws_cam = wb.create_sheet(title=sheet_cambio_turnos)
    ws_cam.append([
        "ID", "Fecha y hora", "LEGAJO", "APELLIDOS Y NOMBRES", "ESPECIALIDAD", "DOTACION",
        "TURNOS", "FRANCO", "LEGAJO2", "APELLIDOS Y NOMBRES2", "ESPECIALIDAD2", "DOTACION2",
        "TURNOS2", "FRANCO2", "Fecha de Cambio de Turno", "REFERENCIA ESTACIÓN", "SUPERVISOR",
        "Observaciones", col_usuario,
    ])

    del wb["Sheet"]
    wb.save(file_path)
